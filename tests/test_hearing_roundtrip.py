from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(ROOT / "scripts" / "hearing"))

import model as M  # noqa: E402
import hearing_sheet as H  # noqa: E402

TEMPLATE = ROOT / "templates" / "sales" / "hearing-sheet.ja.md"


def build_doc() -> dict:
    doc = M.parse_markdown(TEMPLATE.read_text(encoding="utf-8"))
    H.mark_derived(doc, dict(H.DEFAULT_DERIVED))
    H.mark_audience(doc, set(H.DEFAULT_INTERNAL_SECTIONS))
    return doc


class MarkdownRoundTripTest(unittest.TestCase):
    """The Markdown render must survive being read back."""

    def setUp(self) -> None:
        self.doc = build_doc()

    def test_the_template_parses_into_questions_with_unique_ids(self) -> None:
        questions = M.all_questions(self.doc)
        self.assertGreater(len(questions), 100)
        ids = [q["id"] for q in questions]
        self.assertEqual(len(ids), len(set(ids)))
        self.assertEqual([], M.validate(self.doc))

    def test_render_then_parse_keeps_every_answer(self) -> None:
        again = M.parse_markdown(M.render_markdown(self.doc))
        self.assertEqual([q["id"] for q in M.all_questions(self.doc)],
                         [q["id"] for q in M.all_questions(again)])
        self.assertEqual([q["cells"] for q in M.all_questions(self.doc)],
                         [q["cells"] for q in M.all_questions(again)])

    def test_render_is_stable(self) -> None:
        once = M.render_markdown(self.doc)
        twice = M.render_markdown(M.parse_markdown(once))
        self.assertEqual(once, twice)


class XlsxRoundTripTest(unittest.TestCase):
    """An answer typed into the spreadsheet has to come back to the record."""

    def setUp(self) -> None:
        self.doc = build_doc()
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.path = str(Path(self.tmp.name) / "sheet.xlsx")
        import build_sheet
        spec = H.to_sheet_spec(self.doc)
        self.assertEqual([], build_sheet.validate(spec))
        build_sheet.build_xlsx(spec, self.path)

    def _edit(self, qid: str, header: str, value: str) -> None:
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        for ws in wb.worksheets:
            headers = None
            for row in ws.iter_rows(max_row=3, values_only=True):
                if row and row[0] == M.ID_HEADER:
                    headers = list(row)
                    break
            if not headers or header not in headers:
                continue
            column = headers.index(header)
            for row in ws.iter_rows(min_row=2):
                if row[0].value == qid:
                    row[column].value = value
                    wb.save(self.path)
                    return
        self.fail(f"{qid} / {header} が xlsx に見つからない")

    def test_untouched_sheet_reports_no_change(self) -> None:
        incoming = H.read_xlsx_file(self.path)
        changes, conflicts = H.merge(self.doc, incoming, take=None)
        self.assertEqual([], changes)
        self.assertEqual([], conflicts)

    def test_one_edited_answer_comes_back_as_one_change(self) -> None:
        self._edit("4.2-05", "回答", "移行期は直接書き込みが残る")
        incoming = H.read_xlsx_file(self.path)
        changes, conflicts = H.merge(self.doc, incoming, take=None)
        self.assertEqual([], conflicts)
        self.assertEqual(1, len(changes), changes)
        self.assertIn("4.2-05", changes[0])
        block, q = M.find_question(self.doc, "4.2-05")
        self.assertEqual("移行期は直接書き込みが残る", M.get(q, "answer", block["headers"]))

    def test_both_sides_edited_is_a_conflict_and_nothing_is_written(self) -> None:
        baseline = json.loads(json.dumps(self.doc))
        self._edit("4.2-05", "回答", "シート側の回答")
        _, mine = M.find_question(self.doc, "4.2-05")
        mine["cells"]["回答"] = "台帳側の回答"

        incoming = H.read_xlsx_file(self.path)
        changes, conflicts = H.merge(self.doc, incoming, take=None, baseline=baseline)
        self.assertEqual([], changes)
        self.assertEqual(1, len(conflicts))
        self.assertIn("4.2-05", conflicts[0])
        self.assertEqual("台帳側の回答", mine["cells"]["回答"])

    def test_take_sheet_resolves_the_conflict(self) -> None:
        baseline = json.loads(json.dumps(self.doc))
        self._edit("4.2-05", "回答", "シート側の回答")
        _, mine = M.find_question(self.doc, "4.2-05")
        mine["cells"]["回答"] = "台帳側の回答"

        incoming = H.read_xlsx_file(self.path)
        changes, conflicts = H.merge(self.doc, incoming, take="sheet", baseline=baseline)
        self.assertEqual([], conflicts)
        self.assertEqual(1, len(changes))
        self.assertEqual("シート側の回答", mine["cells"]["回答"])


class DerivedSectionTest(unittest.TestCase):
    """The unconfirmed list follows the confidences instead of being kept by hand."""

    def setUp(self) -> None:
        self.doc = build_doc()

    def test_answering_a_question_removes_it_from_the_unconfirmed_list(self) -> None:
        before = M.derived_rows(self.doc, M.DERIVED_UNCONFIRMED)
        block, q = M.find_question(self.doc, "4.2-05")
        q["cells"][M.std_header("confidence", block["headers"])] = "確認済"
        after = M.derived_rows(self.doc, M.DERIVED_UNCONFIRMED)
        self.assertEqual(len(before) - 1, len(after))
        self.assertNotIn("4.2-05", [row[0] for row in after])

    def test_an_estimate_shows_up_in_the_confirm_back_list(self) -> None:
        block, q = M.find_question(self.doc, "6-01")
        q["cells"][M.std_header("confidence", block["headers"])] = "推定"
        q["cells"][M.std_header("answer", block["headers"])] = "突合作業が主な負荷と理解した"
        rows = M.derived_rows(self.doc, M.DERIVED_CONFIRM_BACK)
        self.assertEqual(["6-01"], [row[0] for row in rows])
        self.assertEqual("突合作業が主な負荷と理解した", rows[0][1])

    def test_followup_written_into_the_derived_table_lands_on_the_question(self) -> None:
        M.absorb_derived(self.doc, M.DERIVED_UNCONFIRMED,
                         [["4.2-05", "（無視される）", "情シス部長", "訪問", "6/3", "深津"]])
        _, q = M.find_question(self.doc, "4.2-05")
        self.assertEqual({"target": "情シス部長", "means": "訪問",
                          "due": "6/3", "owner": "深津"}, q["followup"])


class CustomerRenderTest(unittest.TestCase):
    """What is handed to the customer must not carry our own judgements."""

    def setUp(self) -> None:
        self.doc = build_doc()

    def test_internal_columns_and_sections_are_dropped(self) -> None:
        text = M.render_markdown(self.doc, audience="customer")
        self.assertNotIn("| 出典 |", text)
        self.assertNotIn("| 確度 |", text)
        self.assertNotIn("差別化と競合", text)
        self.assertNotIn("金額・BANT", text)
        self.assertIn("現行システムの技術ファクト", text)

    def test_the_dropped_sections_are_reported(self) -> None:
        _, dropped = M.customer_blocks(self.doc)
        self.assertTrue(any("競合" in name for name in dropped))
        self.assertTrue(any("BANT" in name for name in dropped))

    def test_the_customer_spreadsheet_drops_the_same_tabs(self) -> None:
        names = [s["name"] for s in H.to_sheet_spec(self.doc, audience="customer")["sheets"]]
        self.assertFalse(any("競合" in name for name in names))
        self.assertTrue(any("現行システム" in name for name in names))


class GapsTest(unittest.TestCase):
    def test_answered_questions_drop_out_of_gaps(self) -> None:
        doc = build_doc()
        everything = H.gaps(doc)
        block, q = M.find_question(doc, "4.2-05")
        q["cells"][M.std_header("confidence", block["headers"])] = "確認済"
        self.assertEqual(len(everything) - 1, len(H.gaps(doc)))

    def test_section_filter(self) -> None:
        doc = build_doc()
        rows = H.gaps(doc, stage="4")
        self.assertTrue(rows)
        self.assertTrue(all(r["section"].startswith("4") for r in rows))


if __name__ == "__main__":
    unittest.main()
