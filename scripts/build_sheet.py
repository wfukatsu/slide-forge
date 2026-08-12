#!/usr/bin/env python3
"""Build an Excel / Google Spreadsheet from a line-item spec (JSON).

    python scripts/build_sheet.py spec.json --dry-run          # validate only (offline, free)
    python scripts/build_sheet.py spec.json                    # build the xlsx (out/sheets/<title>.xlsx)
    python scripts/build_sheet.py spec.json --out path.xlsx
    python scripts/build_sheet.py spec.json --gsheet [--folder <Drive folder URL/ID>]

Describe a "line items + summary" table -- like a quote, BOM, or cost
breakdown -- in a single JSON file, and assemble it into an xlsx with
openpyxl. Add --gsheet to also convert-upload the same xlsx to Drive as a
Google Spreadsheet (formatting and formulas are converted). The xlsx is
always the source of truth, so both outputs have identical content.

Spec format:

    {
      "title": "Workbook name (becomes the file name / the name in Drive)",
      "sheets": [{
        "name": "Sheet name",
        "heading": "Table heading (optional; shown large on row 1)",
        "note": "Note (optional; e.g. expiration date, tax category)",
        "columns": [
          {"header": "Qty", "type": "int", "width": 8},
          {"header": "Unit price", "type": "currency"},
          {"header": "Amount", "type": "currency", "formula": "=D{row}*F{row}"}
        ],
        "rows": [[...], ...],            // number of columns must match columns; formula columns are null
        "summary": [                      // optional. summary block below the line items
          {"label": "Subtotal",        "formula": "=SUM(G{first}:G{last})"},
          {"label": "Sales tax (10%)", "formula": "=G{s1}*0.1"},
          {"label": "Total",        "formula": "=G{s1}+G{s2}", "emphasis": true}
        ]
      }]
    }

Placeholders in formulas: {row} = the row number of that line item (for
column formulas), {first}/{last} = the first/last row number of the line
items, {s1}, {s2}, ... = the row numbers of summary row 1, row 2, ... (only
earlier summary rows can be referenced).
type: text (default) / int / currency / percent / date.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import _auth  # noqa: E402
from _i18n import t, register  # noqa: E402

register({
    "title is missing": "title がありません",
    "sheets is empty": "sheets が空です",
    "{where}: name is missing": "{where}: name がありません",
    "{where}: columns is empty": "{where}: columns が空です",
    "{where}.columns[{i}]: header is missing": "{where}.columns[{i}]: header がありません",
    "{where}.columns[{i}]: unknown type '{type}' (allowed: {allowed})":
        "{where}.columns[{i}]: 未知の type '{type}'（使えるのは {allowed}）",
    "{where}.columns[{i}]: only {{row}} may be used in a column formula "
    "({{{name}}} was given)":
        "{where}.columns[{i}]: 列 formula で使えるのは {{row}} のみ"
        "（{{{name}}} が指定された）",
    "{where}.rows[{i}]: {got} cells do not match the {want} columns":
        "{where}.rows[{i}]: 列数 {got} が columns の {want} と一致しません",
    "{where}.summary[{i}]: label and formula are required":
        "{where}.summary[{i}]: label と formula が必須です",
    "{where}.summary[{i}]: {{{name}}} refers to a later summary row "
    "(forward references only)":
        "{where}.summary[{i}]: {{{name}}} は自分より後の集計行を参照しています"
        "（前方参照のみ可）",
    "The sheet spec has problems:": "スペックに問題があります:",
    "Validation OK: {sheets} sheet(s) / {rows} line item(s)":
        "検証 OK: {sheets} シート / 明細 {rows} 行",
    "  updated the Google Spreadsheet (existing file with the same name)":
        "  Google Spreadsheet を更新（同名の既存ファイル）",
    "Build an Excel / Google Spreadsheet from a line-item spec":
        "明細表スペックから Excel / Google Spreadsheet を生成する",
    "path to the spec JSON": "スペック JSON のパス",
    "xlsx output path (default: out/sheets/<title>.xlsx)":
        "xlsx の出力パス（省略時: out/sheets/<title>.xlsx）",
    "validate only (no build, no API calls)": "検証のみ（生成しない・API を呼ばない）",
    "also create a Google Spreadsheet (converted upload of the xlsx)":
        "Google Spreadsheet も作成する（xlsx を変換アップロード）",
    "Drive folder URL or ID for the Google Spreadsheet":
        "Google Spreadsheet を置く Drive フォルダの URL または ID",
})

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GSHEET_MIME = "application/vnd.google-apps.spreadsheet"

NUMBER_FORMATS = {
    "text": "@",
    "int": "#,##0",
    "currency": "¥#,##0",
    "percent": "0.0%",
    "date": "yyyy/mm/dd",
}

HEADER_FILL = "1F3864"   # dark navy; a conservative color matched to scalar-2026-family table headers
NOTE_COLOR = "808080"

_PLACEHOLDER_RE = re.compile(r"\{(row|first|last|s\d+)\}")


# ---------- Validation ----------

def validate(spec: dict) -> list[str]:
    errors: list[str] = []
    if not spec.get("title"):
        errors.append(t("title is missing"))
    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        return errors + [t("sheets is empty")]

    for si, sheet in enumerate(sheets):
        where = f"sheets[{si}]"
        if not sheet.get("name"):
            errors.append(t("{where}: name is missing", where=where))
        cols = sheet.get("columns") or []
        if not cols:
            errors.append(t("{where}: columns is empty", where=where))
        for ci, col in enumerate(cols):
            if not col.get("header"):
                errors.append(t("{where}.columns[{i}]: header is missing",
                                where=where, i=ci))
            ctype = col.get("type", "text")
            if ctype not in NUMBER_FORMATS:
                errors.append(t("{where}.columns[{i}]: unknown type '{type}' "
                                "(allowed: {allowed})", where=where, i=ci,
                                type=ctype, allowed="/".join(NUMBER_FORMATS)))
            for m in _PLACEHOLDER_RE.findall(col.get("formula") or ""):
                if m != "row":
                    errors.append(t("{where}.columns[{i}]: only {{row}} may be used "
                                    "in a column formula ({{{name}}} was given)",
                                    where=where, i=ci, name=m))
        for ri, row in enumerate(sheet.get("rows") or []):
            if len(row) != len(cols):
                errors.append(t("{where}.rows[{i}]: {got} cells do not match the "
                                "{want} columns", where=where, i=ri,
                                got=len(row), want=len(cols)))
        for gi, item in enumerate(sheet.get("summary") or []):
            if not item.get("label") or not item.get("formula"):
                errors.append(t("{where}.summary[{i}]: label and formula are required",
                                where=where, i=gi))
                continue
            for m in _PLACEHOLDER_RE.findall(item["formula"]):
                if m.startswith("s") and int(m[1:]) > gi:
                    errors.append(t("{where}.summary[{i}]: {{{name}}} refers to a "
                                    "later summary row (forward references only)",
                                    where=where, i=gi, name=m))
    return errors


# ---------- xlsx generation ----------

def _col_width(col: dict, values: list) -> float:
    if col.get("width"):
        return col["width"]
    # Estimate from measured width, counting CJK characters as width 2
    def w(v) -> int:
        s = str(v) if v is not None else ""
        return sum(2 if ord(ch) > 0x2E7F else 1 for ch in s)
    return min(50, max(8, max([w(col["header"])] + [w(v) for v in values]) + 2))


def build_xlsx(spec: dict, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet in spec["sheets"]:
        ws = wb.create_sheet(sheet["name"])
        cols = sheet["columns"]
        ncols = len(cols)
        r = 1

        if sheet.get("heading"):
            ws.cell(r, 1, sheet["heading"]).font = Font(size=14, bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
        if sheet.get("note"):
            c = ws.cell(r, 1, sheet["note"])
            c.font = Font(size=9, color=NOTE_COLOR)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            r += 1
        if r > 1:
            r += 1  # blank row between the heading block and the table

        header_row = r
        for ci, col in enumerate(cols, 1):
            c = ws.cell(header_row, ci, col["header"])
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.freeze_panes = ws.cell(header_row + 1, 1)

        rows = sheet.get("rows") or []
        first = header_row + 1
        last = header_row + len(rows)
        for ri, row in enumerate(rows):
            rr = first + ri
            for ci, (col, value) in enumerate(zip(cols, row), 1):
                if col.get("formula"):
                    value = col["formula"].replace("{row}", str(rr))
                c = ws.cell(rr, ci, value)
                c.border = border
                c.number_format = NUMBER_FORMATS[col.get("type", "text")]
                if col.get("type") in ("int", "currency", "percent"):
                    c.alignment = Alignment(horizontal="right")

        summary_first = last + 1
        for gi, item in enumerate(sheet.get("summary") or []):
            rr = summary_first + gi
            formula = item["formula"].replace("{first}", str(first)).replace(
                "{last}", str(last))
            for m in set(re.findall(r"\{s(\d+)\}", formula)):
                formula = formula.replace(f"{{s{m}}}", str(summary_first + int(m) - 1))
            label = ws.cell(rr, ncols - 1, item["label"])
            label.font = Font(bold=True)
            label.alignment = Alignment(horizontal="right")
            value = ws.cell(rr, ncols, formula)
            value.number_format = NUMBER_FORMATS[cols[-1].get("type", "currency")]
            value.font = Font(bold=True, size=12 if item.get("emphasis") else 11)
            value.border = border
            if item.get("emphasis"):
                value.fill = PatternFill("solid", fgColor="FFF2CC")

        for ci, col in enumerate(cols, 1):
            values = [row[ci - 1] for row in rows]
            ws.column_dimensions[get_column_letter(ci)].width = _col_width(col, values)

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


# ---------- Google Spreadsheet (convert-upload) ----------

def upload_gsheet(drive, xlsx_path: str, name: str, folder: str | None) -> str:
    """Convert the xlsx into a Google Spreadsheet and create it. If one with the same name already exists, update its contents instead."""
    from googleapiclient.http import MediaFileUpload
    import drive_folder

    media = MediaFileUpload(xlsx_path, mimetype=XLSX_MIME, resumable=False)
    fid = _auth.folder_id(folder) if folder else None
    q = (f"name = '{drive_folder._escape(name)}' and mimeType = '{GSHEET_MIME}' "
         "and trashed = false")
    if fid:
        q += f" and '{fid}' in parents"
    hits = drive.files().list(
        q=q, fields="files(id)", pageSize=5,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute().get("files", [])
    if hits:
        f = drive.files().update(
            fileId=hits[0]["id"], media_body=media, fields="id,webViewLink",
            supportsAllDrives=True,
        ).execute()
        print(t("  updated the Google Spreadsheet (existing file with the same name)"))
    else:
        body: dict = {"name": name, "mimeType": GSHEET_MIME}
        if fid:
            body["parents"] = [fid]
        f = drive.files().create(
            body=body, media_body=media, fields="id,webViewLink",
            supportsAllDrives=True,
        ).execute()
    return f["webViewLink"]


def main() -> int:
    p = argparse.ArgumentParser(
        description=t("Build an Excel / Google Spreadsheet from a line-item spec"))
    p.add_argument("spec", help=t("path to the spec JSON"))
    p.add_argument("--out", help=t("xlsx output path (default: out/sheets/<title>.xlsx)"))
    p.add_argument("--dry-run", action="store_true",
                   help=t("validate only (no build, no API calls)"))
    p.add_argument("--gsheet", action="store_true",
                   help=t("also create a Google Spreadsheet (converted upload of the xlsx)"))
    p.add_argument("--folder",
                   help=t("Drive folder URL or ID for the Google Spreadsheet"))
    args = p.parse_args()

    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)

    errors = validate(spec)
    if errors:
        print(t("The sheet spec has problems:"), file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1
    nrows = sum(len(s.get("rows") or []) for s in spec["sheets"])
    print(t("Validation OK: {sheets} sheet(s) / {rows} line item(s)",
            sheets=len(spec["sheets"]), rows=nrows))
    if args.dry_run:
        return 0

    safe = re.sub(r'[\\/:*?"<>|\s]+', "_", spec["title"]).strip("_")[:80] or "sheet"
    path = args.out or os.path.join("out", "sheets", f"{safe}.xlsx")
    build_xlsx(spec, path)
    print(f"  XLSX: {path}")

    if args.gsheet or args.folder:
        _, drive = _auth.services()
        url = upload_gsheet(drive, path, spec["title"], args.folder)
        print(f"  Google Spreadsheet: {url}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
