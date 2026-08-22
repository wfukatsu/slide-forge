#!/usr/bin/env python3
"""Hearing sheet I/O — one JSON of record, three interchangeable surfaces.

    python scripts/hearing/hearing_sheet.py init templates/sales/hearing-sheet.ja.md \
        --out accounts/<AE>/<customer>/stages/hearing.json
    python scripts/hearing/hearing_sheet.py render hearing.json --format md   --out sheet.md
    python scripts/hearing/hearing_sheet.py render hearing.json --format xlsx --out sheet.xlsx
    python scripts/hearing/hearing_sheet.py render hearing.json --format gsheet --folder <URL>
    python scripts/hearing/hearing_sheet.py read sheet.xlsx --into hearing.json
    python scripts/hearing/hearing_sheet.py gaps hearing.json
    python scripts/hearing/hearing_sheet.py validate hearing.json

The JSON is the record; Markdown, Excel and the Google Spreadsheet are renders
of it. Every question carries a stable ID, so whichever surface someone edited
can be read back and merged.

`read` never overwrites blindly. When the JSON and the surface have both moved
on for the same cell it reports a conflict and stops, unless `--take` says
which side wins — the same discipline as the deal log's contradiction table.

Excel and Google Spreadsheet are produced through `scripts/build_sheet.py`, so
they inherit its validation, its styling and its in-place Drive update (the
URL survives a re-render).
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from typing import Any

HERE = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.dirname(HERE)
ROOT = os.path.dirname(SCRIPTS)
sys.path.insert(0, SCRIPTS)
sys.path.insert(0, HERE)

import model as M  # noqa: E402

CUSTOMER_DROP = {"出典", "確度"}
_SHEET_BAD = re.compile(r"[\[\]:*?/\\]")


# ---------- load / save ----------

def load(path: str) -> dict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def save(doc: dict, path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(doc, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


# ---------- init ----------

DEFAULT_DERIVED = {"12": M.DERIVED_UNCONFIRMED, "13": M.DERIVED_CONFIRM_BACK}

# Sections that must not go on a sheet the customer keeps. Asking these face to
# face is normal; handing over a page that names the opposition, the competitors
# or our own BANT judgement is not (playbook §3). Editable in the JSON.
DEFAULT_INTERNAL_SECTIONS = {"3", "8", "9", "11", "12", "13", "14"}


def mark_audience(doc: dict, internal: set[str]) -> None:
    for block in doc["blocks"]:
        if block["kind"] == "heading" and block.get("level") == 2 and block.get("num"):
            block["customerSafe"] = block["num"] not in internal


def mark_derived(doc: dict, mapping: dict[str, str]) -> int:
    """Turn the hand-kept tables of the named sections into derived blocks.

    The unconfirmed list and the confirm-back list restate what the confidences
    already say. Deriving them removes the second place to keep them in step.
    """
    marked = 0
    num = ""
    for i, block in enumerate(doc["blocks"]):
        if block["kind"] == "heading":
            num = block.get("num") or num
            continue
        if block["kind"] != "prose" or num not in mapping:
            continue
        lines = block["text"].split("\n")
        for j, line in enumerate(lines):
            if line.strip().startswith("|") and j + 1 < len(lines) and \
                    M._ALIGN_RE.match(lines[j + 1]):
                which = mapping[num]
                before = "\n".join(lines[:j]).strip("\n")
                k = j + 2
                while k < len(lines) and lines[k].strip().startswith("|"):
                    k += 1
                after = "\n".join(lines[k:]).strip("\n")
                replacement: list[dict] = []
                if before:
                    replacement.append({"kind": "prose", "text": before})
                replacement.append({"kind": "derived", "which": which,
                                    "headers": M.DERIVED_HEADERS[which],
                                    "align": ["---"] * len(M.DERIVED_HEADERS[which])})
                if after:
                    replacement.append({"kind": "prose", "text": after})
                doc["blocks"][i:i + 1] = replacement
                marked += 1
                break
        if marked and num in mapping:
            mapping = {k: v for k, v in mapping.items() if k != num}
    return marked


def cmd_init(args: argparse.Namespace) -> int:
    with open(args.source, encoding="utf-8") as fh:
        doc = M.parse_markdown(fh.read(), section_prefix=args.section_prefix or "")
    doc["meta"].setdefault("productAddenda", [])
    doc["meta"].setdefault("renders", {})
    if args.product:
        doc["meta"]["productAddenda"] = args.product
    if not args.no_derived:
        mark_derived(doc, dict(DEFAULT_DERIVED))
    mark_audience(doc, set(args.internal_section or DEFAULT_INTERNAL_SECTIONS))
    errors = M.validate(doc)
    if errors:
        print("スペックに問題があります:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1
    save(doc, args.out)
    print(f"作成: {args.out}（設問 {len(M.all_questions(doc))} 件 / "
          f"節 {sum(1 for b in doc['blocks'] if b['kind'] == 'questions')}）")
    return 0


# ---------- render ----------

def sheet_name(num: str, title: str) -> str:
    name = f"{num} {title}" if num else title
    name = _SHEET_BAD.sub("-", name)
    return name[:31] or "sheet"


def to_sheet_spec(doc: dict, *, audience: str = "internal") -> dict:
    """Build the build_sheet.py spec: one tab per question section."""
    fields = (doc.get("meta") or {}).get("fields") or {}
    title = doc.get("title") or "ヒアリングシート"
    sheets: list[dict] = []

    meta_rows = [[k, v] for k, v in fields.items()]
    if meta_rows:
        sheets.append({
            "name": "表紙",
            "heading": title,
            "note": "各タブが 1 節。ID 列は編集しない（この列で JSON と突き合わせる）。",
            "columns": [{"header": "項目", "width": 24}, {"header": "値", "width": 60, "wrap": True}],
            "rows": meta_rows,
        })

    num, heading = "", ""
    blocks = M.customer_blocks(doc)[0] if audience == "customer" else doc["blocks"]
    for block in blocks:
        if block["kind"] == "heading":
            num, heading = block.get("num") or num, block["title"]
            continue
        if block["kind"] != "questions":
            continue
        headers = [h for h in block["headers"] if not (audience == "customer" and h in CUSTOMER_DROP)]
        columns: list[dict] = [{"header": M.ID_HEADER, "width": 9}]
        for h in headers:
            col: dict[str, Any] = {"header": h}
            if h in M.STD["ask"]:
                col.update(width=6, choices=[M.ASK_OFF, M.ASK_ON])
            elif h in M.STD["confidence"]:
                col.update(width=10, choices=list(M.CONFIDENCE))
            elif h in M.STD["audience"]:
                col.update(width=14)
            elif h in M.STD["text"]:
                col.update(width=46, wrap=True)
            elif h in M.STD["answer"]:
                col.update(width=46, wrap=True)
            else:
                col.update(width=22, wrap=True)
            columns.append(col)
        rows = [[q["id"]] + [q["cells"].get(h, "") for h in headers]
                for q in block["questions"]]
        sheets.append({"name": sheet_name(block["num"] or num, heading),
                       "columns": columns, "rows": rows})

    if not sheets:
        raise M.HearingError("設問の節がありません")
    return {"title": title, "sheets": sheets}


def cmd_render(args: argparse.Namespace) -> int:
    doc = load(args.json)
    if args.format == "md":
        text = M.render_markdown(doc, audience=args.audience)
        out = args.out or os.path.join(ROOT, "out", "hearing", "hearing-sheet.md")
        os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
        with open(out, "w", encoding="utf-8") as fh:
            fh.write(text)
        _report_audience(args.audience, doc)
        print(f"出力: {out}")
        if args.audience == "internal":
            doc.setdefault("meta", {}).setdefault("renders", {})["md"] = out
            save(doc, args.json)
        return 0

    import build_sheet

    spec = to_sheet_spec(doc, audience=args.audience)
    errors = build_sheet.validate(spec)
    if errors:
        print("スペックに問題があります:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1
    if args.dry_run:
        print(f"検証 OK: {len(spec['sheets'])} シート / "
              f"設問 {sum(len(s.get('rows') or []) for s in spec['sheets'])} 行")
        return 0

    safe = re.sub(r'[\\/:*?"<>|\s]+', "_", spec["title"]).strip("_")[:80] or "hearing"
    out = args.out or os.path.join(ROOT, "out", "hearing", f"{safe}.xlsx")
    build_sheet.build_xlsx(spec, out)
    _report_audience(args.audience, doc)
    print(f"出力: {out}")
    renders = doc.setdefault("meta", {}).setdefault("renders", {})
    if args.audience == "internal":
        renders["xlsx"] = out

    if args.format == "gsheet":
        import _auth
        _, drive = _auth.services()
        url = build_sheet.upload_gsheet(drive, out, spec["title"], args.folder)
        print(f"Google Spreadsheet: {url}")
        if args.audience == "internal":
            renders["gsheet"] = url
    if args.audience == "internal":
        save(doc, args.json)
    return 0


def _report_audience(audience: str, doc: dict) -> None:
    if audience != "customer":
        return
    columns = sorted({h for b in doc["blocks"] if b["kind"] == "questions"
                      for h in b["headers"] if h in CUSTOMER_DROP})
    _, sections = M.customer_blocks(doc)
    print("顧客配布版として書き出した。落としたもの:")
    print("  列: " + ("、".join(columns) or "なし"))
    print("  節: " + ("、".join(sections) or "なし"))
    print("  これは列と節の機械的な除去にすぎない。"
          "渡す前に本文も目視で確認する（プレイブック §3 の検査）。")


# ---------- read ----------

def read_markdown_file(path: str) -> dict:
    with open(path, encoding="utf-8") as fh:
        return M.parse_markdown(fh.read())


def read_xlsx_file(path: str) -> dict:
    """Read back an xlsx render: one tab per section, ID in the first column."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    doc: dict = {"schemaVersion": M.SCHEMA_VERSION, "meta": {}, "blocks": []}
    for ws in wb.worksheets:
        header_row, headers = None, []
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
            values = [("" if c is None else str(c).strip()) for c in row]
            if values and values[0] == M.ID_HEADER:
                header_row = values
                headers = [h for h in values[1:] if h]
                break
        if not header_row:
            continue
        start = next(i for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1)
            if [("" if c is None else str(c).strip()) for c in row][:1] == [M.ID_HEADER]) + 1
        questions = []
        for row in ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True):
            values = [("" if c is None else str(c).strip()) for c in row]
            if not values or not values[0]:
                continue
            cells = {h: (values[i + 1] if i + 1 < len(values) else "")
                     for i, h in enumerate(headers)}
            questions.append({"id": values[0], "cells": cells,
                              "followup": {}, "confirmBack": {}})
        if questions:
            doc["blocks"].append({"kind": "questions", "num": ws.title.split(" ")[0],
                                  "headers": headers, "align": ["---"] * len(headers),
                                  "questions": questions})
    return doc


def read_gsheet(url_or_id: str, tmp_dir: str) -> dict:
    """Export the Google Spreadsheet as xlsx via Drive, then read it as an xlsx.

    Uses the Drive scope the toolchain already holds — no Sheets scope is added,
    so nobody has to re-authenticate.
    """
    import _auth
    import drive_folder

    _, drive = _auth.services()
    path = os.path.join(tmp_dir, "hearing-export.xlsx")
    drive_folder.download(drive, drive_folder.file_id(url_or_id), path)
    return read_xlsx_file(path)


def merge(doc: dict, incoming: dict, *, take: str | None,
          baseline: dict | None = None) -> tuple[list[str], list[str]]:
    """Merge the edited surface into the record. Returns (changes, conflicts)."""
    changes: list[str] = []
    conflicts: list[str] = []
    base_by_id = {}
    if baseline:
        for block, q in M._questions(baseline):
            base_by_id[q["id"]] = q["cells"]

    for in_block, in_q in M._questions(incoming):
        block, q = M.find_question(doc, in_q["id"])
        if q is None:
            conflicts.append(f"{in_q['id']}: 台帳に無い ID（行が足されたか、ID が書き換えられた）")
            continue
        for header, value in in_q["cells"].items():
            if header not in block["headers"]:
                continue
            current = q["cells"].get(header, "")
            if value == current:
                continue
            base = base_by_id.get(in_q["id"], {}).get(header)
            both_moved = base is not None and base != current and base != value
            if both_moved and take is None:
                conflicts.append(
                    f"{in_q['id']} / {header}: 双方が変更されている"
                    f"（台帳「{current}」/ 取り込み側「{value}」）")
                continue
            if both_moved and take == "json":
                continue
            q["cells"][header] = value
            changes.append(f"{in_q['id']} / {header}: 「{current}」→「{value}」")

    for block in incoming.get("blocks", []):
        if block["kind"] == "derived":
            M.absorb_derived(doc, block["which"], block.get("rows") or [])
    return changes, conflicts


def cmd_read(args: argparse.Namespace) -> int:
    doc = load(args.into)
    src = args.source
    if src.startswith("http") or "/spreadsheets/" in src:
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            incoming = read_gsheet(src, tmp)
    elif src.endswith(".xlsx"):
        incoming = read_xlsx_file(src)
    elif src.endswith(".md"):
        incoming = read_markdown_file(src)
    else:
        print(f"読めない形式です: {src}（.md / .xlsx / Google Spreadsheet の URL）",
              file=sys.stderr)
        return 1

    baseline = load(args.baseline) if args.baseline else None
    changes, conflicts = merge(doc, incoming, take=args.take, baseline=baseline)

    for line in changes:
        print(f"  変更 {line}")
    if conflicts:
        print("\n競合（取り込みを中止した）:", file=sys.stderr)
        for line in conflicts:
            print(f"  - {line}", file=sys.stderr)
        print("\n--take sheet / --take json で、どちらを採るかを明示してください。",
              file=sys.stderr)
        return 1
    if args.dry_run:
        print(f"\n（--dry-run のため書き込んでいない） 変更 {len(changes)} 件")
        return 0
    save(doc, args.into)
    print(f"\n取り込み: {args.into}（変更 {len(changes)} 件）")
    return 0


# ---------- gaps ----------

def gaps(doc: dict, *, stage: str | None = None) -> list[dict]:
    out = []
    num, heading = "", ""
    for block in doc["blocks"]:
        if block["kind"] == "heading":
            num, heading = block.get("num") or num, block["title"]
            continue
        if block["kind"] != "questions":
            continue
        if stage and not (block["num"] or num).startswith(stage):
            continue
        for q in block["questions"]:
            confidence = M.get(q, "confidence", block["headers"])
            if confidence == "確認済":
                continue
            out.append({
                "id": q["id"],
                "section": block["num"] or num,
                "sectionTitle": heading,
                "text": M.get(q, "text", block["headers"]),
                "answer": M.get(q, "answer", block["headers"]),
                "audience": M.get(q, "audience", block["headers"]),
                "confidence": confidence or "未確認",
                "ask": M.get(q, "ask", block["headers"]) == M.ASK_ON,
                "followup": q.get("followup") or {},
            })
    return out


def cmd_gaps(args: argparse.Namespace) -> int:
    doc = load(args.json)
    rows = gaps(doc, stage=args.section)
    if args.json_out:
        print(json.dumps(rows, ensure_ascii=False, indent=2))
        return 0
    if not rows:
        print("未確認・推定はありません。")
        return 0
    estimated = [r for r in rows if r["confidence"] == "推定"]
    print(f"未確認 {len(rows) - len(estimated)} 件 / 推定 {len(estimated)} 件\n")
    section = None
    for r in rows:
        if r["section"] != section:
            section = r["section"]
            print(f"§{section} {r['sectionTitle']}")
        mark = "★" if r["ask"] else " "
        print(f"  {mark} [{r['confidence']}] {r['id']}  {r['text']}")
    print("\n★ = 「聞く」に印が付いている設問。"
          "スライドにするときは、この順で hearing-slides に渡す。")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    doc = load(args.json)
    errors = M.validate(doc)
    if errors:
        print("問題があります:", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        return 1
    print(f"検証 OK: 設問 {len(M.all_questions(doc))} 件 / "
          f"節 {sum(1 for b in doc['blocks'] if b['kind'] == 'questions')} / "
          f"派生節 {sum(1 for b in doc['blocks'] if b['kind'] == 'derived')}")
    return 0


# ---------- CLI ----------

def main() -> int:
    p = argparse.ArgumentParser(description="ヒアリングシートの入出力（JSON が正本）")
    sub = p.add_subparsers(dest="cmd", required=True)

    q = sub.add_parser("init", help="Markdown の様式から JSON を起こす")
    q.add_argument("source", help="templates/sales/hearing-sheet.ja.md など")
    q.add_argument("--out", required=True)
    q.add_argument("--product", action="append", help="使う製品補遺の名前（複数可）")
    q.add_argument("--section-prefix", help="補遺を取り込むときの ID 接頭辞（例: scalar-）")
    q.add_argument("--no-derived", action="store_true",
                   help="未確認リスト・確認を返すことを派生節にしない")
    q.add_argument("--internal-section", action="append",
                   help="顧客配布版から落とす節番号（既定: "
                        + " ".join(sorted(DEFAULT_INTERNAL_SECTIONS)) + "）")
    q.set_defaults(func=cmd_init)

    q = sub.add_parser("render", help="md / xlsx / gsheet へ出力する")
    q.add_argument("json")
    q.add_argument("--format", choices=["md", "xlsx", "gsheet"], default="md")
    q.add_argument("--out")
    q.add_argument("--folder", help="Google Spreadsheet を置く Drive フォルダ")
    q.add_argument("--audience", choices=["internal", "customer"], default="internal")
    q.add_argument("--dry-run", action="store_true", help="検証のみ（API を呼ばない）")
    q.set_defaults(func=cmd_render)

    q = sub.add_parser("read", help="編集された md / xlsx / Google Spreadsheet を取り込む")
    q.add_argument("source")
    q.add_argument("--into", required=True)
    q.add_argument("--baseline", help="前回書き出した JSON。双方の変更を検出するのに使う")
    q.add_argument("--take", choices=["sheet", "json"], help="競合時にどちらを採るか")
    q.add_argument("--dry-run", action="store_true")
    q.set_defaults(func=cmd_read)

    q = sub.add_parser("gaps", help="未確認・推定の一覧")
    q.add_argument("json")
    q.add_argument("--section", help="節番号で絞る（例: 4）")
    q.add_argument("--json-out", action="store_true")
    q.set_defaults(func=cmd_gaps)

    q = sub.add_parser("validate", help="語彙・ID・派生節を検査する")
    q.add_argument("json")
    q.set_defaults(func=cmd_validate)

    args = p.parse_args()
    try:
        return args.func(args)
    except M.HearingError as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
